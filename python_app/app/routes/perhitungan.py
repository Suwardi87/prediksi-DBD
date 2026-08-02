"""
Perhitungan Manual Routes — Random Forest
163 data, 4 fitur (X1=Usia, X2=LamaRawat, X3=JK, X4=JumlahKasus)
15 pohon, setiap pohon HANYA 1 fitur untuk split.
Encoding: Rendah=1, Sedang=2, Tinggi=3
Grouping (Bab IV Tabel 4.2): Kasus 1-10=Rendah, 11-20=Sedang, >20=Tinggi
Pohon 11 (Sampel 11) = pohon terbaik (MAE terkecil). Evaluasi pada 10 data uji.
"""
import os
import math
import numpy as np
from flask import Blueprint, render_template, jsonify
from flask_login import login_required

perhitungan_bp = Blueprint('perhitungan', __name__)

EXCEL_PATH = os.path.normpath(os.path.join(
    os.path.dirname(__file__), '..', '..', '..', 'Data DBD 15 Sampel All.xlsx'
))

LABEL_ENCODE = {'Rendah': 1, 'Sedang': 2, 'Tinggi': 3}
LABEL_DECODE = {1: 'Rendah', 2: 'Sedang', 3: 'Tinggi'}

COL_MAP = {
    'Usia': 'usia',
    'Lama Rawat Inap': 'lama_rawat',
    'Jumlah Kasus Per Bulan': 'jumlah_kasus',
    'Jumlah Kasus Perbulan': 'jumlah_kasus',
    'Jenis Kelamin': 'jk',
    'Tingkat Resiko': 'tingkat_risiko',
    'Terdeteksi': 'tingkat_risiko',
    'Nama': 'nama',
}

POHON_NAMES = [
    'Pohon 1', 'Pohon 2', 'Pohon 3', 'Pohon 4', 'Pohon 5',
    'Pohon 6',     'Pohon 7', 'Pohon 8', 'Pohon 9', 'Pohon 10',
    'Pohon 11', 'Pohon 12', 'Pohon 13', 'Pohon 14', 'Pohon 15',
]

PEMILIHAN_FITUR_HEADERS = ['Fitur', 'Nilai']

N_TEST = 10




def _parse_numeric(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().lower().replace('hari', '').strip()
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0

def _read_data_dbd(wb):
    if 'Data_DBD' not in wb.sheetnames:
        return []
    ws = wb['Data_DBD']
    seen_cols = set()
    headers = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h and h in COL_MAP and COL_MAP[h] not in seen_cols:
            headers.append((c, COL_MAP[h]))
            seen_cols.add(COL_MAP[h])
    data = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, feat_name in headers:
            row[feat_name] = ws.cell(row=r, column=c).value
        if row.get('tingkat_risiko') in LABEL_ENCODE:
            data.append(row)
        else:
            tr = row.get('tingkat_risiko')
            if isinstance(tr, (int, float)) and LABEL_DECODE.get(int(tr)):
                row['tingkat_risiko'] = LABEL_DECODE[int(tr)]
                data.append(row)
    return data


def _read_bootstrap_from_sheet(wb, pohon_name):
    if pohon_name not in wb.sheetnames:
        return []
    ws = wb[pohon_name]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    nos = [i for i, h in enumerate(headers) if h == 'No']
    rs = nos[1] if len(nos) >= 2 else None
    if rs is None:
        return []

    right_headers = []
    for i in range(rs, len(headers)):
        h = headers[i]
        if h and h in COL_MAP:
            right_headers.append((i, COL_MAP[h]))
        elif h == 'Perhitungan root':
            break

    samples = []
    for r in range(2, ws.max_row + 1):
        no_val = ws.cell(row=r, column=rs + 1).value
        if no_val is None or not isinstance(no_val, (int, float)):
            continue
        sample = {}
        for col_idx, feat_name in right_headers:
            raw = ws.cell(row=r, column=col_idx + 1).value
            if feat_name in ('usia', 'lama_rawat', 'jk', 'jumlah_kasus'):
                sample[feat_name] = _parse_numeric(raw)
            else:
                sample[feat_name] = raw
        tr = sample.get('tingkat_risiko')
        if tr is not None:
            if isinstance(tr, (int, float)):
                tr = LABEL_DECODE.get(int(tr))
                if tr:
                    sample['tingkat_risiko'] = tr
            if sample.get('tingkat_risiko') in LABEL_ENCODE:
                samples.append(sample)
    return samples


def _read_excel_entropy(wb, pohon_name):
    ws = wb[pohon_name]
    for r in range(1, min(ws.max_row + 1, 100)):
        for c in range(1, min(ws.max_column + 1, 100)):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str) and 'entropy root' in v.lower():
                return ws.cell(row=r, column=c + 1).value
    return None


def _find_calc_header_row(ws, label, col_start=15, max_row=300):
    for r in range(1, min(ws.max_row + 1, max_row)):
        v = ws.cell(row=r, column=col_start).value
        if v and isinstance(v, str) and label.lower() in v.lower():
            return r
    return None


def _read_pohon_root_calc(wb, pohon_name):
    ws = wb[pohon_name]
    hdr_row = _find_calc_header_row(ws, 'Perhitungan root')
    if hdr_row is None:
        return None
    ent_row = _find_calc_header_row(ws, 'Entropy Root')
    if ent_row is None:
        return None
    counts = {}
    probs = {}
    for r in range(hdr_row + 1, ent_row):
        lbl = ws.cell(row=r, column=15).value
        cnt = ws.cell(row=r, column=16).value
        prb = ws.cell(row=r, column=17).value
        if lbl is not None and cnt is not None:
            key = int(lbl) if isinstance(lbl, (int, float)) else str(lbl)
            counts[key] = int(cnt) if isinstance(cnt, (int, float)) else 0
            probs[key] = float(prb) if isinstance(prb, (int, float)) else 0.0
    root_entropy = ws.cell(row=ent_row, column=16).value
    return {
        'class_counts': counts,
        'class_probs': probs,
        'root_entropy': float(root_entropy) if isinstance(root_entropy, (int, float)) else 0.0,
    }


def _read_pohon_thresholds(wb, pohon_name):
    ws = wb[pohon_name]
    hdr_row = _find_calc_header_row(ws, 'JUMLAH KASUS')
    if hdr_row is None:
        hdr_row = _find_calc_header_row(ws, 'Jumlah kasus')
    if hdr_row is None:
        return None, None
    t1 = ws.cell(row=hdr_row + 1, column=16).value
    t2 = ws.cell(row=hdr_row + 1, column=17).value
    t1 = float(t1) if isinstance(t1, (int, float)) else None
    t2 = float(t2) if isinstance(t2, (int, float)) else None
    return t1, t2


def _read_pohon_gain(wb, pohon_name):
    ws = wb[pohon_name]
    for r in range(150, min(ws.max_row + 1, 250)):
        for c in range(15, 18):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str) and v.strip().upper() == 'IG':
                ig = ws.cell(row=r + 1, column=c).value
                ea = ws.cell(row=r + 1, column=c + 1).value
                if isinstance(ig, (int, float)) and isinstance(ea, (int, float)):
                    return float(ig), float(ea)
    return None, None


def _read_test_actuals(wb):
    if 'Data Uji' not in wb.sheetnames:
        return []
    ws = wb['Data Uji']
    data = []
    for r in range(2, ws.max_row + 1):
        jk = ws.cell(row=r, column=3).value
        risk = ws.cell(row=r, column=5).value
        if jk is not None and risk is not None:
            data.append({
                'jumlah_kasus': int(jk) if isinstance(jk, (int, float)) else 0,
                'risk_enc': int(risk) if isinstance(risk, (int, float)) else 0,
            })
    return data


def _read_pohon_features(wb, pohon_name):
    ws = wb[pohon_name]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    nos = [i for i, h in enumerate(headers) if h == 'No']
    rs = nos[1] if len(nos) >= 2 else None
    if rs is None:
        return []
    features = []
    for i in range(rs, len(headers)):
        h = headers[i]
        if h and h in COL_MAP and COL_MAP[h] != 'tingkat_risiko':
            features.append(COL_MAP[h])
        elif h == 'Perhitungan root':
            break
    return features


def _calc_entropy(counts):
    total = sum(counts)
    if total == 0:
        return 0.0
    ent = 0.0
    for c in counts:
        if c > 0:
            p = c / total
            ent -= p * math.log2(p)
    return ent


def _calc_root_entropy(samples):
    counts = {'Rendah': 0, 'Sedang': 0, 'Tinggi': 0}
    for s in samples:
        r = s.get('tingkat_risiko', '')
        if r in counts:
            counts[r] += 1
    return _calc_entropy([counts['Rendah'], counts['Sedang'], counts['Tinggi']]), counts


BINARY_FEATURES = {'jk'}


def _compute_split_with_thresholds(samples, feature_key, t1, t2):
    if not samples:
        return None, None, 0.0, None, None, None

    root_e, root_counts = _calc_root_entropy(samples)
    n = len(samples)

    left = [s for s in samples if float(s.get(feature_key, 0)) < t1]
    mid = [s for s in samples if t1 <= float(s.get(feature_key, 0)) <= t2]
    right = [s for s in samples if float(s.get(feature_key, 0)) > t2]

    left_e = _calc_root_entropy(left)[0] if left else 0
    mid_e = _calc_root_entropy(mid)[0] if mid else 0
    right_e = _calc_root_entropy(right)[0] if right else 0
    weighted_e = (len(left) / n) * left_e + (len(mid) / n) * mid_e + (len(right) / n) * right_e
    gain = root_e - weighted_e

    return t1, t2, gain, root_e, root_counts, {
        'left_entropy': round(left_e, 6),
        'mid_entropy': round(mid_e, 6),
        'right_entropy': round(right_e, 6),
        'left_samples': len(left),
        'mid_samples': len(mid),
        'right_samples': len(right),
        'weighted_entropy': round(weighted_e, 6),
    }


def _find_best_split_single_feature(samples, feature_key):
    if not samples:
        return None, None, 0.0, None, None, None

    root_e, root_counts = _calc_root_entropy(samples)
    n = len(samples)
    values = sorted(set(float(s.get(feature_key, 0)) for s in samples))

    if len(values) <= 1:
        return None, None, 0.0, root_e, root_counts, None

    if feature_key in BINARY_FEATURES:
        t1 = 0.5
        t2 = 0.5
        left = [s for s in samples if float(s.get(feature_key, 0)) < t1]
        mid = [s for s in samples if t1 <= float(s.get(feature_key, 0)) <= t2]
        right = [s for s in samples if float(s.get(feature_key, 0)) > t2]
        left_e = _calc_root_entropy(left)[0] if left else 0
        mid_e = _calc_root_entropy(mid)[0] if mid else 0
        right_e = _calc_root_entropy(right)[0] if right else 0
        weighted_e = (len(left) / n) * left_e + (len(mid) / n) * mid_e + (len(right) / n) * right_e
        gain = root_e - weighted_e
        return t1, t2, gain, root_e, root_counts, {
            'left_entropy': round(left_e, 6),
            'mid_entropy': round(mid_e, 6),
            'right_entropy': round(right_e, 6),
            'left_samples': len(left),
            'mid_samples': len(mid),
            'right_samples': len(right),
            'weighted_entropy': round(weighted_e, 6),
        }

    midpoints = [(values[i] + values[i + 1]) / 2.0 for i in range(len(values) - 1)]

    best_gain = -1
    best_t1 = None
    best_t2 = None
    best_left = None
    best_mid = None
    best_right = None

    for i in range(len(midpoints)):
        for j in range(i + 1, len(midpoints)):
            t1 = midpoints[i]
            t2 = midpoints[j]
            left = [s for s in samples if float(s.get(feature_key, 0)) < t1]
            mid = [s for s in samples if t1 <= float(s.get(feature_key, 0)) <= t2]
            right = [s for s in samples if float(s.get(feature_key, 0)) > t2]
            if not left or not right:
                continue

            left_e = _calc_root_entropy(left)[0] if left else 0
            mid_e = _calc_root_entropy(mid)[0] if mid else 0
            right_e = _calc_root_entropy(right)[0] if right else 0
            weighted_e = (len(left) / n) * left_e + (len(mid) / n) * mid_e + (len(right) / n) * right_e
            gain = root_e - weighted_e

            if gain > best_gain:
                best_gain = gain
                best_t1 = t1
                best_t2 = t2
                best_left = left
                best_mid = mid
                best_right = right

    if best_t1 is None:
        return None, None, 0.0, root_e, root_counts, None

    left_e = _calc_root_entropy(best_left)[0] if best_left else 0
    mid_e = _calc_root_entropy(best_mid)[0] if best_mid else 0
    right_e = _calc_root_entropy(best_right)[0] if best_right else 0
    weighted_e = (len(best_left) / n) * left_e + (len(best_mid) / n) * mid_e + (len(best_right) / n) * right_e

    return best_t1, best_t2, best_gain, root_e, root_counts, {
        'left_entropy': round(left_e, 6),
        'mid_entropy': round(mid_e, 6),
        'right_entropy': round(right_e, 6),
        'left_samples': len(best_left),
        'mid_samples': len(best_mid),
        'right_samples': len(best_right),
        'weighted_entropy': round(weighted_e, 6),
    }


FEATURE_NAMES = {
    'usia': 'Usia',
    'lama_rawat': 'Lama Rawat Inap',
    'jk': 'Jenis Kelamin',
    'jumlah_kasus': 'Jumlah Kasus Perbulan',
}

FEATURE_DISPLAY = {
    'usia': 'Usia (X1)',
    'lama_rawat': 'Lama Rawat Inap (X2)',
    'jk': 'Jenis Kelamin (X3)',
    'jumlah_kasus': 'Jumlah Kasus Perbulan (X4)',
}


def _majority_class(subset):
    counts = {'Rendah': 0, 'Sedang': 0, 'Tinggi': 0}
    for s in subset:
        r = s.get('tingkat_risiko', '')
        if r in counts:
            counts[r] += 1
    return max(counts, key=counts.get), counts


def _build_tree_rules_deep(samples, feature_key, depth=0, max_depth=3, fixed_t1=None, fixed_t2=None):
    if not samples or depth >= max_depth:
        return []

    root_e, root_counts = _calc_root_entropy(samples)
    n = len(samples)

    all_same = len([c for c in root_counts.values() if c > 0]) <= 1
    if all_same or n <= 5:
        cls, cls_counts = _majority_class(samples)
        return [{'type': 'leaf', 'class': cls, 'counts': cls_counts, 'n': n}]

    if fixed_t1 is not None and fixed_t2 is not None and depth == 0:
        t1, t2 = fixed_t1, fixed_t2
        _, _, gain, _, _, _ = _compute_split_with_thresholds(samples, feature_key, t1, t2)
    else:
        t1, t2, gain, re_tmp, rc, info = _find_best_split_single_feature(samples, feature_key)
    if t1 is None or gain <= 0:
        cls, cls_counts = _majority_class(samples)
        return [{'type': 'leaf', 'class': cls, 'counts': cls_counts, 'n': n}]

    fname = FEATURE_NAMES.get(feature_key, feature_key)
    left = [s for s in samples if float(s.get(feature_key, 0)) < t1]
    mid = [s for s in samples if t1 <= float(s.get(feature_key, 0)) <= t2]
    right = [s for s in samples if float(s.get(feature_key, 0)) > t2]

    left_majority, left_counts = _majority_class(left) if left else ('Rendah', {'Rendah': 0, 'Sedang': 0, 'Tinggi': 0})
    mid_majority, mid_counts = _majority_class(mid) if mid else ('Sedang', {'Rendah': 0, 'Sedang': 0, 'Tinggi': 0})
    right_majority, right_counts = _majority_class(right) if right else ('Tinggi', {'Rendah': 0, 'Sedang': 0, 'Tinggi': 0})

    result = [{
        'type': 'split',
        'feature': fname,
        'threshold_low': round(t1, 2),
        'threshold_high': round(t2, 2),
        'gain': round(gain, 4),
        'left_class': left_majority,
        'left_counts': left_counts,
        'left_n': len(left),
        'mid_class': mid_majority,
        'mid_counts': mid_counts,
        'mid_n': len(mid),
        'right_class': right_majority,
        'right_counts': right_counts,
        'right_n': len(right),
    }]

    if left and len(left) > 2:
        result.extend(_build_tree_rules_deep(left, feature_key, depth + 1, max_depth))
    if mid and len(mid) > 2:
        result.extend(_build_tree_rules_deep(mid, feature_key, depth + 1, max_depth))
    if right and len(right) > 2:
        result.extend(_build_tree_rules_deep(right, feature_key, depth + 1, max_depth))

    return result


def _build_rules_text(samples, feature_key, fixed_t1=None, fixed_t2=None):
    if not samples:
        return []

    if fixed_t1 is not None and fixed_t2 is not None:
        t1, t2 = fixed_t1, fixed_t2
    else:
        t1, t2, gain, re, rc, info = _find_best_split_single_feature(samples, feature_key)
    if t1 is None:
        cls, counts = _majority_class(samples)
        return [f"IF {FEATURE_NAMES.get(feature_key, feature_key)} ANY THEN Risiko = {cls} (n={len(samples)}, {counts})"]

    fname = FEATURE_NAMES.get(feature_key, feature_key)
    left = [s for s in samples if float(s.get(feature_key, 0)) < t1]
    mid = [s for s in samples if t1 <= float(s.get(feature_key, 0)) <= t2]
    right = [s for s in samples if float(s.get(feature_key, 0)) > t2]

    left_cls, left_counts = _majority_class(left) if left else ('Rendah', {'Rendah': 0, 'Sedang': 0, 'Tinggi': 0})
    mid_cls, mid_counts = _majority_class(mid) if mid else ('Sedang', {'Rendah': 0, 'Sedang': 0, 'Tinggi': 0})
    right_cls, right_counts = _majority_class(right) if right else ('Tinggi', {'Rendah': 0, 'Sedang': 0, 'Tinggi': 0})

    rules = []
    rules.append(f"IF {fname} < {t1:.2f} THEN Risiko = {left_cls} (n={len(left)}, {left_counts})")
    rules.append(f"IF {fname} >= {t1:.2f} AND <= {t2:.2f} THEN Risiko = {mid_cls} (n={len(mid)}, {mid_counts})")
    rules.append(f"IF {fname} > {t2:.2f} THEN Risiko = {right_cls} (n={len(right)}, {right_counts})")
    return rules


def _read_pemilihan_fitur(wb):
    if 'Pemilihan Fitur' not in wb.sheetnames:
        return []
    ws = wb['Pemilihan Fitur']
    data = []
    for r in range(3, ws.max_row + 1):
        fitur = ws.cell(row=r, column=1).value
        nilai = ws.cell(row=r, column=2).value
        if fitur:
            data.append({
                'fitur': str(fitur),
                'nilai': round(float(nilai), 6) if nilai and isinstance(nilai, (int, float)) else str(nilai or ''),
            })
    return data


def _read_perhitungan_rf(wb):
    if 'PerhitunganRF' not in wb.sheetnames:
        return []
    ws = wb['PerhitunganRF']
    trees = []
    current = None
    for r in range(1, ws.max_row + 1):
        c1 = ws.cell(row=r, column=1).value
        c8 = ws.cell(row=r, column=8).value
        c9 = ws.cell(row=r, column=9).value

        if c1 and isinstance(c1, str) and c1.startswith('Sampel Pohon'):
            if current:
                trees.append(current)
            current = {'name': c1, 'data_rows': [], 'metrics': {}}
        elif current and c8 and c9:
            metric_key = str(c8).strip()
            try:
                current['metrics'][metric_key] = round(float(c9), 6)
            except (ValueError, TypeError):
                pass
    if current:
        trees.append(current)
    return trees


def _read_penentuan_pohon_terbaik(wb):
    if 'PenentuanPohonterbaik' not in wb.sheetnames:
        return []
    ws = wb['PenentuanPohonterbaik']
    data = []
    for r in range(2, ws.max_row + 1):
        no = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=2).value
        mae = ws.cell(row=r, column=3).value
        rmse = ws.cell(row=r, column=4).value
        r2 = ws.cell(row=r, column=5).value
        if no is not None:
            data.append({
                'no': int(no) if isinstance(no, (int, float)) else no,
                'name': str(name or ''),
                'mae': round(float(mae), 4) if mae else None,
                'rmse': round(float(rmse), 4) if rmse else None,
                'r2': round(float(r2), 4) if r2 else None,
            })
    return data


def _read_perhitungan_data_uji(wb):
    if 'perhitungan data uji' not in wb.sheetnames:
        return None
    ws = wb['perhitungan data uji']
    metrics = {}
    data_rows = []
    for r in range(2, ws.max_row + 1):
        c8 = ws.cell(row=r, column=8).value
        c9 = ws.cell(row=r, column=9).value
        if c8 and c9:
            try:
                metrics[str(c8).strip()] = round(float(c9), 6)
            except (ValueError, TypeError):
                pass
    return metrics


def _predict_with_thresholds(jumlah_kasus, thresholds):
    if jumlah_kasus < thresholds[0]:
        return 'Rendah'
    elif jumlah_kasus <= thresholds[1]:
        return 'Sedang'
    else:
        return 'Tinggi'


def _read_sheet_table(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return {'headers': [], 'rows': []}
    ws = wb[sheet_name]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = []
        valid = False
        for c in range(1, len(headers) + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                valid = True
            row.append(str(v) if v is not None else '')
        if valid:
            rows.append(row)
    return {'headers': headers, 'rows': rows}


@perhitungan_bp.route('/')
@login_required
def index():
    return render_template('perhitungan/index.html')


@perhitungan_bp.route('/hitung', methods=['POST'])
@login_required
def hitung():
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

        data_dbd = _read_data_dbd(wb)
        total = len(data_dbd)

        label_counts = {'Rendah': 0, 'Sedang': 0, 'Tinggi': 0}
        for d in data_dbd:
            lr = d.get('tingkat_risiko')
            if lr in label_counts:
                label_counts[lr] += 1

        data_uji_excel = _read_sheet_table(wb, 'Data Uji')
        data_latih_excel = _read_sheet_table(wb, 'Data Latih')

        step1_data = []
        for i, d in enumerate(data_dbd):
            step1_data.append({
                'no': i + 1,
                'nama': d.get('nama', ''),
                'usia': _parse_numeric(d.get('usia')),
                'lama_rawat': _parse_numeric(d.get('lama_rawat')),
                'jk': _parse_numeric(d.get('jk')),
                'jk_label': 'L' if d.get('jk') == 1 else 'P',
                'jumlah_kasus': float(d.get('jumlah_kasus') or 0),
                'tingkat_risiko': d.get('tingkat_risiko', ''),
                'tingkat_risiko_encoded': LABEL_ENCODE.get(d.get('tingkat_risiko'), 0),
            })

        test_data = data_dbd[:N_TEST]
        train_data = data_dbd[N_TEST:]

        pemilihan_fitur_data = _read_pemilihan_fitur(wb)
        perhitungan_rf_data = _read_perhitungan_rf(wb)
        penentuan_pohon_terbaik_data = _read_penentuan_pohon_terbaik(wb)
        perhitungan_data_uji_metrics = _read_perhitungan_data_uji(wb)

        pohon_results = []

        for pohon_idx, pohon_name in enumerate(POHON_NAMES):
            pohon_num = pohon_idx + 1
            bootstrap_samples = _read_bootstrap_from_sheet(wb, pohon_name)
            if not bootstrap_samples:
                continue

            pohon_features = _read_pohon_features(wb, pohon_name)
            excel_entropy = _read_excel_entropy(wb, pohon_name)

            bab4_t1, bab4_t2 = _read_pohon_thresholds(wb, pohon_name)
            bab4_feat = 'jumlah_kasus'

            if bab4_t1 is not None and bab4_t2 is not None:
                best_threshold_low = bab4_t1
                best_threshold_high = bab4_t2
                best_feature = bab4_feat
                _, _, best_gain, _, best_root_counts, best_split_info = _compute_split_with_thresholds(
                    bootstrap_samples, bab4_feat, bab4_t1, bab4_t2
                )
                if best_split_info is None:
                    best_gain = 0
                    best_split_info = {'weighted_entropy': 0, 'left_entropy': 0, 'mid_entropy': 0,
                                       'right_entropy': 0, 'left_samples': 0, 'mid_samples': 0, 'right_samples': 0}
                    best_root_counts = {'Rendah': 0, 'Sedang': 0, 'Tinggi': 0}
            else:
                best_threshold_low = None
                best_threshold_high = None
                best_gain = -1
                best_feature = None
                best_split_info = None
                best_root_counts = None

                for feat in pohon_features:
                    t1, t2, gain, root_e, root_counts, split_info = _find_best_split_single_feature(
                        bootstrap_samples, feat
                    )
                    if gain > best_gain:
                        best_gain = gain
                        best_threshold_low = t1
                        best_threshold_high = t2
                        best_feature = feat
                        best_split_info = split_info
                        best_root_counts = root_counts

            root_entropy = best_split_info['weighted_entropy'] + best_gain if best_split_info else 0

            rules_text = _build_rules_text(bootstrap_samples, best_feature, best_threshold_low, best_threshold_high)
            rules_deep = _build_tree_rules_deep(bootstrap_samples, best_feature, 0, 3, best_threshold_low, best_threshold_high)

            unique_hashes = set()
            for s in bootstrap_samples:
                unique_hashes.add(str(sorted(s.items())))
            n_unique = len(unique_hashes)

            bab4_gain, bab4_entropy_after = _read_pohon_gain(wb, pohon_name)
            if bab4_gain is None:
                bab4_gain = best_gain
            if bab4_entropy_after is None:
                bab4_entropy_after = best_split_info['weighted_entropy'] if best_split_info else 0

            display_gain = bab4_gain
            display_root_entropy = round(bab4_gain + bab4_entropy_after, 6)

            pohon_results.append({
                'name': pohon_name,
                'id': pohon_num,
                'target_feature': best_feature,
                'target_label': FEATURE_NAMES.get(best_feature, best_feature),
                'target_display': FEATURE_DISPLAY.get(best_feature, best_feature),
                'n_samples': len(bootstrap_samples),
                'n_unique': n_unique,
                'n_duplicates': len(bootstrap_samples) - n_unique,
                'features_available': pohon_features,
                'best_feature': best_feature,
                'threshold_low': round(best_threshold_low, 2) if best_threshold_low is not None else None,
                'threshold_high': round(best_threshold_high, 2) if best_threshold_high is not None else None,
                'gain': round(display_gain, 6),
                'gain_bab4': round(bab4_gain, 6),
                'root_entropy': display_root_entropy,
                'root_counts': {'Tinggi': best_root_counts.get('Tinggi', 0),
                                'Sedang': best_root_counts.get('Sedang', 0),
                                'Rendah': best_root_counts.get('Rendah', 0)} if best_root_counts else {},
                'left_entropy': best_split_info.get('left_entropy', 0) if best_split_info else 0,
                'mid_entropy': best_split_info.get('mid_entropy', 0) if best_split_info else 0,
                'right_entropy': best_split_info.get('right_entropy', 0) if best_split_info else 0,
                'left_samples': best_split_info.get('left_samples', 0) if best_split_info else 0,
                'mid_samples': best_split_info.get('mid_samples', 0) if best_split_info else 0,
                'right_samples': best_split_info.get('right_samples', 0) if best_split_info else 0,
                'weighted_entropy': best_split_info.get('weighted_entropy', 0) if best_split_info else 0,
                'entropy_bab4': round(bab4_entropy_after, 6),
                'rules_text': rules_text,
                'rules_deep': rules_deep,
                'excel_entropy': round(excel_entropy, 6) if excel_entropy is not None else None,
            })

        penentuan_data = _read_penentuan_pohon_terbaik(wb)
        if penentuan_data:
            best_entry = max(penentuan_data, key=lambda x: x.get('r2') or -999999)
            best_tree_idx = best_entry.get('no', 6) - 1
        else:
            best_tree_idx = 5
        best_tree = pohon_results[best_tree_idx] if best_tree_idx < len(pohon_results) else None

        excel_test = _read_test_actuals(wb)
        p11_t1, p11_t2 = _read_pohon_thresholds(wb, POHON_NAMES[best_tree_idx])
        p11_bootstrap = _read_bootstrap_from_sheet(wb, POHON_NAMES[best_tree_idx])
        p11_left = [s for s in p11_bootstrap if float(s.get('jumlah_kasus', 0)) < p11_t1]
        p11_mid = [s for s in p11_bootstrap if p11_t1 <= float(s.get('jumlah_kasus', 0)) <= p11_t2]
        p11_right = [s for s in p11_bootstrap if float(s.get('jumlah_kasus', 0)) > p11_t2]
        p11_left_cls, _ = _majority_class(p11_left) if p11_left else ('Sedang', {})
        p11_mid_cls, _ = _majority_class(p11_mid) if p11_mid else ('Tinggi', {})
        p11_right_cls, _ = _majority_class(p11_right) if p11_right else ('Tinggi', {})

        bab4_test_results = []
        for i, td in enumerate(excel_test):
            jk = td['jumlah_kasus']
            actual_enc = td['risk_enc']
            actual = LABEL_DECODE.get(actual_enc, 'Sedang')
            if jk < p11_t1:
                predicted = p11_left_cls
            elif jk <= p11_t2:
                predicted = p11_mid_cls
            else:
                predicted = p11_right_cls
            predicted_enc = LABEL_ENCODE.get(predicted, 2)
            bab4_test_results.append({
                'no': i + 1,
                'jumlah_kasus': jk,
                'actual': actual,
                'actual_enc': actual_enc,
                'predicted': predicted,
                'predicted_enc': predicted_enc,
                'correct': actual == predicted,
            })

        bab4_correct = sum(1 for t in bab4_test_results if t['correct'])
        bab4_accuracy = round(bab4_correct / N_TEST, 4)

        bab4_actual_enc = [t['actual_enc'] for t in bab4_test_results]
        bab4_pred_enc = [t['predicted_enc'] for t in bab4_test_results]
        bab4_actual = np.array(bab4_actual_enc, dtype=float)
        bab4_pred = np.array(bab4_pred_enc, dtype=float)
        bab4_abs_err = np.abs(bab4_actual - bab4_pred)
        bab4_sq_err = (bab4_actual - bab4_pred) ** 2
        bab4_mae = round(float(np.mean(bab4_abs_err)), 4)
        bab4_rmse = round(float(math.sqrt(np.mean(bab4_sq_err))), 4)
        bab4_ss_res = float(np.sum(bab4_sq_err))
        bab4_ss_tot = float(np.sum((bab4_actual - 1) ** 2))
        bab4_r2 = round(1 - bab4_ss_res / bab4_ss_tot, 4) if bab4_ss_tot > 0 else 0.0

        our_test_results = []
        our_threshold_used = None
        if best_tree:
            thr_low = best_tree.get('threshold_low')
            thr_high = best_tree.get('threshold_high')
            left_class, mid_class, right_class = 'Rendah', 'Sedang', 'Tinggi'
            if best_tree.get('rules_deep'):
                for rule in best_tree['rules_deep']:
                    if rule.get('type') == 'split':
                        left_class = rule.get('left_class', 'Rendah')
                        mid_class = rule.get('mid_class', 'Sedang')
                        right_class = rule.get('right_class', 'Tinggi')
                        break
            if thr_low is not None and thr_high is not None:
                our_threshold_used = f'{thr_low}, {thr_high}'
                target_feat = best_tree.get('target_feature', 'jumlah_kasus') if best_tree else 'jumlah_kasus'
                for i in range(N_TEST):
                    actual_risk = test_data[i].get('tingkat_risiko', '')
                    feature_val = _parse_numeric(test_data[i].get(target_feat, 0))
                    if feature_val < thr_low:
                        predicted_risk = left_class
                    elif feature_val <= thr_high:
                        predicted_risk = mid_class
                    else:
                        predicted_risk = right_class
                    our_test_results.append({
                        'no': i + 1,
                        'feature_name': FEATURE_NAMES.get(target_feat, target_feat),
                        'feature_val': feature_val,
                        'actual': actual_risk,
                        'actual_enc': LABEL_ENCODE.get(actual_risk, 0),
                        'predicted': predicted_risk,
                        'predicted_enc': LABEL_ENCODE.get(predicted_risk, 0),
                        'correct': actual_risk == predicted_risk,
                    })

        our_correct = sum(1 for t in our_test_results if t['correct'])
        our_accuracy = round(our_correct / N_TEST, 4)

        our_actual_enc = [t['actual_enc'] for t in our_test_results]
        our_pred_enc = [t['predicted_enc'] for t in our_test_results]
        our_actual = np.array(our_actual_enc, dtype=float)
        our_pred = np.array(our_pred_enc, dtype=float)
        our_abs_err = np.abs(our_actual - our_pred)
        our_sq_err = (our_actual - our_pred) ** 2
        our_mae = round(float(np.mean(our_abs_err)), 4)
        our_rmse = round(float(math.sqrt(np.mean(our_sq_err))), 4)
        our_y_mean = float(np.mean(our_actual))
        our_ss_res = float(np.sum(our_sq_err))
        our_ss_tot = float(np.sum((our_actual - our_y_mean) ** 2))
        our_r2 = round(1 - our_ss_res / our_ss_tot, 4) if our_ss_tot > 0 else 0.0

        return jsonify({
            'status': 'success',
            'step1': {
                'data': step1_data,
                'total': total,
                'n_train': len(train_data),
                'n_test': N_TEST,
                'features': ['Usia (X1)', 'Lama Rawat Inap (X2)', 'Jenis Kelamin (X3)', 'Jumlah Kasus Perbulan (X4)'],
                'label_counts': label_counts,
                'encoding': {'Rendah': 1, 'Sedang': 2, 'Tinggi': 3},
            },
            'data_uji': data_uji_excel,
            'data_latih': data_latih_excel,
            'pemilihan_fitur': pemilihan_fitur_data,
            'step2': {
                'encoding_table': [
                    {'fitur': 'Jenis Kelamin (X3)', 'nilai_asli': 'Laki-laki (L) / Perempuan (P)', 'nilai_encoding': '1 / 0'},
                    {'fitur': 'Tingkat Resiko (Target)', 'nilai_asli': 'Rendah / Sedang / Tinggi', 'nilai_encoding': '1 / 2 / 3'},
                ],
                'grouping_table': [
                    {'fitur': 'Jumlah Kasus', 'rentang': '1 – 10', 'risiko': 'Rendah', 'label': 1},
                    {'fitur': 'Jumlah Kasus', 'rentang': '11 – 20', 'risiko': 'Sedang', 'label': 2},
                    {'fitur': 'Jumlah Kasus', 'rentang': '> 20', 'risiko': 'Tinggi', 'label': 3},
                ],
            },
            'step4': {
                'trees': pohon_results,
            },
            'perhitungan_rf': perhitungan_rf_data,
            'penentuan_pohon_terbaik': penentuan_pohon_terbaik_data,
            'perhitungan_data_uji_metrics': perhitungan_data_uji_metrics,
            'step5': {
                'best_tree': best_tree,
                'bab4_rules': _build_rules_text(p11_bootstrap, 'jumlah_kasus', p11_t1, p11_t2),
                'bab4_thresholds': [round(p11_t1, 2), round(p11_t2, 2)],
                'best_tree_name': f'Pohon {best_tree_idx + 1} (Sampel {best_tree_idx + 1})',
                'bab4_test_data': bab4_test_results,
                'bab4_correct': bab4_correct,
                'bab4_accuracy': bab4_accuracy,
                'bab4_mae': bab4_mae,
                'bab4_rmse': bab4_rmse,
                'bab4_r2': bab4_r2,
                'our_test_data': our_test_results,
                'our_correct': our_correct,
                'our_accuracy': our_accuracy,
                'our_mae': our_mae,
                'our_rmse': our_rmse,
                'our_r2': our_r2,
                'our_threshold': our_threshold_used,
                'our_rules': best_tree.get('rules_text', []) if best_tree else [],
            },
            'step6': {
                'bab4_mae': bab4_mae,
                'bab4_rmse': bab4_rmse,
                'bab4_r2': bab4_r2,
                'bab4_accuracy': bab4_accuracy,
                'bab4_correct': bab4_correct,
                'our_mae': our_mae,
                'our_rmse': our_rmse,
                'our_r2': our_r2,
                'our_accuracy': our_accuracy,
                'our_correct': our_correct,
                'our_threshold': our_threshold_used,
                'n_test': N_TEST,
            },
        })

    except Exception as e:
        import traceback
        return jsonify({
            'status': 'error',
            'message': 'Terjadi kesalahan saat menghitung perhitungan manual.',
        }), 500
